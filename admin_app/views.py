# admin_app/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags, format_html
from django.db import transaction
import logging
import re
import os
from django.conf import settings
from django.core.mail import EmailMessage
from django.shortcuts import render

from university_app.models import UniversityRequest
from admin_app.models import (
    ApprovedUniversityAccount,
    ApprovedUniversityCredential
)

# ----------------------------------
# LOGGER
# ----------------------------------
logger = logging.getLogger(__name__)

# ----------------------------------
# ADMIN PANEL URL
# ----------------------------------
ADMIN_URL = "http://127.0.0.1:8000/admin/"

# ----------------------------------
# HOME PAGE
# ----------------------------------
def home(request):
    context = {
        "page_title": "Welcome to SmartEduCare",
        "current_year": 2025,
        "featured_message": "Register your university today!",
    }
    return render(request, "index.html", context)

# ----------------------------------
# INDEX PAGE
# ----------------------------------
def indexs(request):
    context = {
        "page_title": "Welcome to SmartEduCare",
        "current_year": 2025,
        "featured_message": "Register your university today!",
    }
    return render(request, "indexs.html", context)

# ----------------------------------
# SHOW UNIVERSITY REQUESTS
# ----------------------------------
def university_reg_show(request):
    universities = UniversityRequest.objects.all().order_by("-created_at")
    return render(request, "university_reg_show.html", {"universities": universities})

# ----------------------------------
# REJECT UNIVERSITY
# ----------------------------------
def admin_reject(request, pk):
    uni = get_object_or_404(UniversityRequest, pk=pk)

    if uni.status != "Pending":
        messages.warning(request, "Only pending requests can be rejected.")
        return redirect("university_reg_show")

    uni.status = "Rejected"
    uni.save()

    messages.success(request, f'University "{uni.name}" has been rejected.')
    return redirect("university_reg_show")

from django.shortcuts import render, redirect
from django.contrib import messages

def admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if username == "1005" and password == "Shariff@0307":
            request.session['admin_logged_in'] = True
            return redirect('admin_dash')
        else:
            messages.error(request, "Invalid Username or Password")

    return render(request, "admin_login.html")

import os
import glob
import pandas as pd
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages


# ─────────────────────────────────────────────────────────────────────────────
#  Directory where uploaded files are stored
# ─────────────────────────────────────────────────────────────────────────────
RECEIVED_DIR = os.path.join(
    settings.BASE_DIR,
    "static",
    "data",
    "uploads",
    "university_data",
)


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _scan_files():
    """Return { display_label: full_path } for every Excel/CSV in RECEIVED_DIR."""
    # Ensure directory exists
    os.makedirs(RECEIVED_DIR, exist_ok=True)
    
    file_map = {}
    for ext in ("*.csv", "*.xls", "*.xlsx"):
        for fp in glob.glob(os.path.join(RECEIVED_DIR, ext)):
            label = os.path.basename(fp)
            file_map[label] = fp
    return file_map


def _read_file(file_path: str) -> pd.DataFrame:
    """Read CSV or Excel into a DataFrame."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        try:
            return pd.read_csv(file_path, dtype=str)
        except UnicodeDecodeError:
            return pd.read_csv(file_path, dtype=str, encoding="latin-1")
    else:
        return pd.read_excel(file_path, dtype=str)


def _save_file(df: pd.DataFrame, file_path: str):
    """Write DataFrame back to the original file (same format)."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        df.to_csv(file_path, index=False)
    else:
        df.to_excel(file_path, index=False)


def _find_roll_col(df: pd.DataFrame):
    """
    Heuristic: return the column whose name contains roll/id keywords.
    Falls back to the first column.
    """
    priority_keywords = [
        "roll", "rollno", "roll_no", "rollno.",
        "studentid", "student_id", "studentid.",
        "regno", "reg_no", "id",
    ]
    # Normalised → original column map
    lower_cols = {
        c.lower().replace(" ", "").replace("_", "").replace(".", ""): c
        for c in df.columns
    }
    for kw in priority_keywords:
        if kw in lower_cols:
            return lower_cols[kw]
    for col in df.columns:
        cl = col.lower()
        if "roll" in cl or ("student" in cl and "id" in cl):
            return col
    return df.columns[0] if len(df.columns) > 0 else None


def _validate_file_extension(filename: str) -> bool:
    """Check if file has a valid extension (.csv, .xls, .xlsx)."""
    ext = os.path.splitext(filename)[1].lower()
    return ext in (".csv", ".xls", ".xlsx")


def _get_safe_filename(filename: str) -> str:
    """
    Sanitize filename to prevent path traversal and other security issues.
    Removes any directory separators and keeps only the base filename.
    """
    # Get just the basename (removes any path components)
    safe_name = os.path.basename(filename)
    
    # Remove any remaining potentially dangerous characters
    # Keep only alphanumeric, dots, underscores, hyphens, and spaces
    safe_chars = []
    for char in safe_name:
        if char.isalnum() or char in '._- ':
            safe_chars.append(char)
        else:
            safe_chars.append('_')
    
    return ''.join(safe_chars)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN VIEW
# ─────────────────────────────────────────────────────────────────────────────

def admin_dash(request):

    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    file_map     = _scan_files()
    universities = sorted(file_map.keys())   # ["file_a.xlsx", "file_b.csv", …]

    # ── context defaults ──────────────────────────────────────────────────────
    selected_uni = ""
    roll_no      = ""
    show_all     = ""
    search_done  = False
    student_data = []
    headers      = []
    col_pairs    = []
    all_data     = []
    all_headers  = []

    # ══════════════════════════════════════════════════════════════════════════
    # POST — two sub-actions:
    #   1) "upload"  — user uploads one or more files from their PC
    #   2) (default) — modify column(s) for a student and save back
    # ══════════════════════════════════════════════════════════════════════════
    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        # ── 1. UPLOAD FILES ────────────────────────────────────────────────
        if action == "upload":
            # Ensure the directory exists before uploading
            os.makedirs(RECEIVED_DIR, exist_ok=True)
            
            uploaded_files = request.FILES.getlist("files")
            
            if not uploaded_files:
                messages.error(request, "No files were selected for upload.")
            else:
                saved, skipped, errors = [], [], []
                
                for f in uploaded_files:
                    # Validate file extension
                    if not _validate_file_extension(f.name):
                        skipped.append(f.name)
                        continue
                    
                    try:
                        # Sanitize filename for security
                        safe_filename = _get_safe_filename(f.name)
                        
                        # Construct full destination path
                        dest_path = os.path.join(RECEIVED_DIR, safe_filename)
                        
                        # Check if file already exists (optional: you can auto-rename)
                        if os.path.exists(dest_path):
                            messages.warning(
                                request,
                                f"File '{safe_filename}' already exists and will be overwritten."
                            )
                        
                        # Write file to disk
                        with open(dest_path, "wb+") as dest_fp:
                            for chunk in f.chunks():
                                dest_fp.write(chunk)
                        
                        saved.append(safe_filename)
                        
                    except Exception as e:
                        errors.append(f"{f.name} (Error: {str(e)})")

                # Display success messages
                if saved:
                    messages.success(
                        request,
                        f"✅ Uploaded {len(saved)} file(s) to 'static/data/uploads/university_data/': {', '.join(saved)}"
                    )
                
                # Display warnings for skipped files
                if skipped:
                    messages.warning(
                        request,
                        f"⚠️ Skipped {len(skipped)} unsupported file(s) (only .csv, .xls, .xlsx allowed): {', '.join(skipped)}"
                    )
                
                # Display errors
                if errors:
                    messages.error(
                        request,
                        f"❌ Failed to upload {len(errors)} file(s): {', '.join(errors)}"
                    )

            return redirect("/admin_dash/")

        # ── 2. MODIFY STUDENT RECORD ───────────────────────────────────────
        selected_uni = request.POST.get("university", "").strip()
        roll_no      = request.POST.get("roll_no",    "").strip()
        file_path    = file_map.get(selected_uni)

        if not file_path:
            messages.error(request, f"File '{selected_uni}' not found.")
        else:
            try:
                df       = _read_file(file_path)
                roll_col = _find_roll_col(df)

                if not roll_col:
                    messages.error(request, "Could not identify the roll-number column.")
                else:
                    mask = df[roll_col].astype(str).str.strip() == roll_no

                    if not mask.any():
                        messages.error(
                            request,
                            f"Roll No. '{roll_no}' not found in '{selected_uni}'."
                        )
                    else:
                        updated = []
                        for key, new_val in request.POST.items():
                            if not key.startswith("col_"):
                                continue
                            col_name = key[4:]        # strip "col_" prefix
                            new_val  = new_val.strip()
                            if not new_val:
                                continue
                            matched = next(
                                (c for c in df.columns if c.strip() == col_name),
                                None
                            )
                            if matched:
                                df.loc[mask, matched] = new_val
                                updated.append(f"{matched} → {new_val}")

                        if updated:
                            _save_file(df, file_path)
                            messages.success(
                                request,
                                f"✅ Updated {len(updated)} field(s) for '{roll_no}': "
                                + " | ".join(updated)
                            )
                        else:
                            messages.error(request, "No changes were submitted.")

            except Exception as e:
                messages.error(request, f"Error updating file: {e}")

        return redirect(
            f"/admin_dash/?university={selected_uni}&roll_no={roll_no}"
        )

    # ══════════════════════════════════════════════════════════════════════════
    # GET — search / show-all
    # ══════════════════════════════════════════════════════════════════════════
    # Re-scan so newly uploaded files appear immediately
    file_map     = _scan_files()
    universities = sorted(file_map.keys())

    selected_uni = request.GET.get("university", "").strip()
    roll_no      = request.GET.get("roll_no",    "").strip()
    show_all     = request.GET.get("show_all",   "").strip()

    if selected_uni:
        file_path = file_map.get(selected_uni)

        if not file_path:
            messages.error(
                request,
                f"File '{selected_uni}' not found. "
                "Make sure it exists in static/data/uploads/verify/"
            )
        else:
            try:
                df = _read_file(file_path)

                # ── show all students ─────────────────────────────────────
                if show_all:
                    all_headers = list(df.columns)
                    all_data    = df.fillna("").values.tolist()

                # ── search specific student ───────────────────────────────
                if roll_no:
                    search_done = True
                    roll_col    = _find_roll_col(df)

                    if roll_col:
                        result = df[
                            df[roll_col].astype(str).str.strip() == roll_no
                        ]
                        if not result.empty:
                            headers      = list(result.columns)
                            student_data = result.fillna("").values.tolist()
                            first_row    = result.fillna("").iloc[0]
                            col_pairs    = [
                                (str(col).strip(), str(val))
                                for col, val in zip(result.columns, first_row.tolist())
                            ]
                    else:
                        messages.error(
                            request, "Could not identify the roll-number column."
                        )

            except Exception as e:
                messages.error(request, f"Error reading file: {e}")

    context = {
        "universities": universities,
        "selected_uni": selected_uni,
        "roll_no":      roll_no,
        "show_all":     show_all,
        "search_done":  search_done,
        "student_data": student_data,
        "headers":      headers,
        "col_pairs":    col_pairs,
        "all_data":     all_data,
        "all_headers":  all_headers,
        "upload_path":  "static/data/uploads/university_data/",  # For display in template
    }
    return render(request, "admin_dash.html", context)
# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _read_file(fp):
    ext = os.path.splitext(fp)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(fp, dtype=str)
    elif ext == ".csv":
        try:
            df = pd.read_csv(fp, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(fp, dtype=str, encoding="latin-1")
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    df.columns = df.columns.str.strip()
    return df


def _save_file(df, fp):
    ext = os.path.splitext(fp)[1].lower()
    if ext in (".xlsx", ".xls"):
        df.to_excel(fp, index=False)
    elif ext == ".csv":
        df.to_csv(fp, index=False, encoding="utf-8-sig")
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _find_roll_col(df):
    candidates = {
        "rollno","roll_no","rollnumber","roll_number",
        "studentid","student_id","id","regno","reg_no",
        "enrollment","enrollmentno","admissionnumber","admission_no","studentrollno",
    }
    for col in df.columns:
        n = col.strip().lower().replace(" ","").replace("_","").replace("-","")
        if n in candidates:
            return col
    return df.columns[0] if len(df.columns) > 0 else None


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _read_file(file_path):
    """
    Load a file into a DataFrame.
    Supports: .xlsx, .xls  → pd.read_excel
              .csv          → pd.read_csv  (tries utf-8, falls back to latin-1)
    All values read as str to avoid type issues.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(file_path, dtype=str)

    elif ext == ".csv":
        try:
            df = pd.read_csv(file_path, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(file_path, dtype=str, encoding="latin-1")

    else:
        raise ValueError(f"Unsupported file type: {ext}")

    df.columns = df.columns.str.strip()
    return df


def _save_file(df, file_path):
    """
    Save DataFrame back to the original file format.
    .xlsx / .xls → to_excel (no index)
    .csv          → to_csv  (no index, utf-8-sig for Excel-safe CSV)
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        df.to_excel(file_path, index=False)

    elif ext == ".csv":
        df.to_csv(file_path, index=False, encoding="utf-8-sig")

    else:
        raise ValueError(f"Unsupported file type for saving: {ext}")


def _find_roll_col(df):
    """
    Detect the roll-number / student-ID column by matching common names.
    Falls back to the first column if nothing matches.
    """
    candidates = {
        "rollno", "roll_no", "rollnumber", "roll_number",
        "studentid", "student_id", "id", "regno",
        "reg_no", "enrollment", "enrollmentno",
        "admissionnumber", "admission_no", "studentrollno",
    }
    for col in df.columns:
        normalised = col.strip().lower().replace(" ", "").replace("_", "").replace("-", "")
        if normalised in candidates:
            return col
    return df.columns[0] if len(df.columns) > 0 else None


def _find_col(df, keywords):
    """
    Find the first column whose name contains any of the given keywords
    (case-insensitive substring match). Returns column name or None.
    """
    for col in df.columns:
        col_lower = col.strip().lower()
        if any(kw in col_lower for kw in keywords):
            return col
    return None

import os
import re
import logging
from datetime import datetime

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.mail import EmailMultiAlternatives, EmailMessage
from django.core.files.storage import FileSystemStorage
from django.template.loader import render_to_string
from django.utils.html import strip_tags, format_html
from django.db import transaction
from django.conf import settings

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import ApprovedUniversityCredential
from student_app.models import student_login_valid

logger = logging.getLogger(__name__)

ADMIN_URL = "/admin/"

# ─────────────────────────────────────────────────────────────────────────────
# BASE PATHS
# ─────────────────────────────────────────────────────────────────────────────
REF_EXCEL_PATH = os.path.join(
    settings.BASE_DIR,
    "static", "data", "uploads", "REF_Share",
    "students_final_v8.xlsx",
)

# Root folder — all password-generated outputs live here
PW_GEN_ROOT = os.path.join(
    settings.BASE_DIR,
    "static", "data", "uploads", "password_generated_files",
)

LOGINS_DIR  = os.path.join(
    settings.BASE_DIR,
    "static", "data", "uploads", "univesity_logins",
)
LOGINS_FILE = os.path.join(LOGINS_DIR, "university_logins.xlsx")

AISHE_DIR = os.path.join(
    settings.BASE_DIR,
    "static", "data", "uploads", "AISHE_CODE",
)

LOGIN_HEADERS = [
    "S.No", "College ID / Username", "Official Email",
    "Password", "University Name", "Approved On",
]

AISHE_COPY_COLUMNS = ["StudentID", "AISHE_Code", "student_password", "College_name"]


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — sanitise a string so it is safe as a folder / file name
# ─────────────────────────────────────────────────────────────────────────────
def _safe_dirname(s: str) -> str:
    s = re.sub(r'[\\/:*?"<>|,\s]+', "_", s.strip())
    s = re.sub(r"_+", "_", s)
    return s.strip("_") or "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — derive & create the per-university output folder
#
# Path:  PW_GEN_ROOT / {UniversityName}_{AISHE_Code} /
#
# • Folder is created  if it does not exist.
# • If it already exists it is REUSED  (files inside are overwritten).
# ─────────────────────────────────────────────────────────────────────────────
def _get_university_folder(university_name: str, aishe_code: str) -> str:
    dir_name    = _safe_dirname(f"{university_name}_{aishe_code}")
    folder_path = os.path.join(PW_GEN_ROOT, dir_name)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — append one credential row to university_logins.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def _save_credentials_to_excel(college_id, email, password, name):
    os.makedirs(LOGINS_DIR, exist_ok=True)
    file_exists = os.path.isfile(LOGINS_FILE)

    if file_exists:
        wb       = openpyxl.load_workbook(LOGINS_FILE)
        ws       = wb.active
        next_row = ws.max_row + 1
        serial   = next_row - 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "University Logins"

        hdr_fill   = PatternFill("solid", fgColor="0C1828")
        hdr_font   = Font(name="Calibri", bold=True, color="00E5A0", size=11)
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hdr_border = Border(
            left=Side(style="thin", color="1A3050"),
            right=Side(style="thin", color="1A3050"),
            top=Side(style="thin", color="1A3050"),
            bottom=Side(style="thin", color="1A3050"),
        )
        for ci, txt in enumerate(LOGIN_HEADERS, 1):
            c           = ws.cell(row=1, column=ci, value=txt)
            c.fill      = hdr_fill
            c.font      = hdr_font
            c.alignment = hdr_align
            c.border    = hdr_border

        for i, w in enumerate([8, 28, 34, 28, 36, 22], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        next_row = 2
        serial   = 1

    row_data    = [serial, college_id, email, password, name,
                   datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    row_fill    = PatternFill("solid", fgColor="0C1828" if serial % 2 == 0 else "111F30")
    cell_border = Border(
        left=Side(style="thin", color="162438"), right=Side(style="thin", color="162438"),
        top=Side(style="thin", color="162438"),  bottom=Side(style="thin", color="162438"),
    )
    col_fonts = {
        1: Font(name="Calibri",     size=10, color="6B7A94", bold=True),
        2: Font(name="Courier New", size=10, color="4FACFE"),
        3: Font(name="Calibri",     size=10, color="D8E2F0"),
        4: Font(name="Courier New", size=10, color="00E5A0"),
        5: Font(name="Calibri",     size=10, color="D8E2F0", bold=True),
        6: Font(name="Calibri",     size=9,  color="6B7A94", italic=True),
    }
    col_aligns = {1: Alignment(horizontal="center", vertical="center")}

    for ci, val in enumerate(row_data, 1):
        c           = ws.cell(row=next_row, column=ci, value=val)
        c.fill      = row_fill
        c.font      = col_fonts.get(ci, Font(name="Calibri", size=10, color="D8E2F0"))
        c.alignment = col_aligns.get(ci, Alignment(horizontal="left", vertical="center"))
        c.border    = cell_border

    ws.row_dimensions[next_row].height = 22
    wb.save(LOGINS_FILE)
    logger.info("Credentials appended row=%s college_id=%s", next_row, college_id)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — save AISHE_CODE copy
# ─────────────────────────────────────────────────────────────────────────────
def save_aishe_copy(df: pd.DataFrame) -> list:
    os.makedirs(AISHE_DIR, exist_ok=True)
    missing = [c for c in ["StudentID", "AISHE_Code", "student_password"]
               if c not in df.columns]
    if missing:
        logger.warning("AISHE copy skipped — missing: %s", missing)
        return []

    copy_df = df.copy()
    copy_df = copy_df.dropna(subset=["StudentID", "AISHE_Code"])
    copy_df = copy_df[copy_df["StudentID"].astype(str).str.strip() != ""]
    copy_df = copy_df[copy_df["AISHE_Code"].astype(str).str.strip() != ""]
    for col in ["StudentID", "AISHE_Code", "student_password"]:
        copy_df[col] = copy_df[col].astype(str).str.strip()

    def _college_name(code):
        try:
            return ApprovedUniversityCredential.objects.get(college_id=code).name.strip()
        except ApprovedUniversityCredential.DoesNotExist:
            return code

    copy_df["College_name"] = copy_df["AISHE_Code"].apply(_college_name)
    copy_df = copy_df[AISHE_COPY_COLUMNS]

    saved = []
    for code, grp in copy_df.groupby("AISHE_Code"):
        out = os.path.join(AISHE_DIR, f"{code}.xlsx")
        if os.path.exists(out):
            existing = pd.read_excel(out, dtype=str).fillna("")
            merged   = pd.concat([existing, grp], ignore_index=True)
            merged   = merged.drop_duplicates(subset=["StudentID"], keep="last")
            merged.to_excel(out, index=False)
        else:
            grp.reset_index(drop=True).to_excel(out, index=False)
        saved.append(f"{code}.xlsx")
    return saved


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — process ONE uploaded file object
#   • Generates student_password = StudentID + PAN_CARD[-4:]
#   • Saves into   PW_GEN_ROOT / {UniversityName}_{AISHE} / password_generated_{fname}
#   • Runs AISHE copy
#   Returns a result dict  { filename, output_name, relative_folder,
#                            row_count, aishe_files, error }
# ─────────────────────────────────────────────────────────────────────────────
def _process_single_file(file_obj, university_name: str, aishe_code: str) -> dict:
    fname = file_obj.name
    ext   = os.path.splitext(fname)[1].lower()

    # Read
    try:
        if ext == ".csv":
            df, is_csv = pd.read_csv(file_obj, dtype=str), True
        else:
            df, is_csv = pd.read_excel(file_obj, dtype=str), False
        df = df.fillna("")
    except Exception as exc:
        return {"filename": fname, "error": f"Cannot read file: {exc}"}

    # Validate
    required = {"StudentID", "PAN_CARD"}
    if not required.issubset(df.columns):
        missing = required - set(df.columns)
        return {"filename": fname,
                "error": f"Missing required columns: {', '.join(sorted(missing))}"}

    # Generate password
    df["student_password"] = (
        df["StudentID"].astype(str).str.strip()
        + df["PAN_CARD"].astype(str).str.strip().str[-4:]
    )

    # Resolve output folder:  PW_GEN_ROOT / UniversityName_AISHE /
    uni_folder  = _get_university_folder(university_name, aishe_code)
    output_name = f"password_generated_{fname}"
    output_path = os.path.join(uni_folder, output_name)

    # Save
    try:
        if is_csv:
            df.to_csv(output_path, index=False)
        else:
            df.to_excel(output_path, index=False)
    except Exception as exc:
        return {"filename": fname, "error": f"Cannot save output: {exc}"}

    # AISHE copy
    aishe_files = []
    if "AISHE_Code" in df.columns:
        aishe_files = save_aishe_copy(df)

    return {
        "filename":        fname,
        "output_name":     output_name,
        "output_path":     output_path,
        "relative_folder": os.path.relpath(uni_folder, settings.BASE_DIR),
        "row_count":       len(df),
        "aishe_files":     aishe_files,
        "error":           None,
    }


# ═════════════════════════════════════════════════════════════════════════════
#  APPROVE UNIVERSITY  (unchanged logic)
# ═════════════════════════════════════════════════════════════════════════════
def admin_approve(request, pk):
    uni = get_object_or_404(UniversityRequest, pk=pk)
    if uni.status != "Pending":
        messages.warning(request, "Only pending requests can be approved.")
        return redirect("university_reg_show")

    clean_college_id = re.sub(r"[,!_\-=*&^%$#@]", "", str(uni.college_id))
    clean_pincode    = re.sub(r"[,!_\-=*&^%$#@]", "", str(uni.pincode))
    raw_password     = f"Smartedu_{clean_college_id}_{clean_pincode}"
    email_name       = re.sub(r"\d+", "", uni.university_email.split("@")[0]).capitalize()

    context = {
        "clean_name": email_name,
        "college_id": uni.college_id,
        "email":      uni.university_email,
        "password":   raw_password,
        "login_url":  request.build_absolute_uri("/login/"),
    }

    try:
        with transaction.atomic():
            if ApprovedUniversityAccount.objects.filter(college_id=uni.college_id).exists():
                messages.error(request, format_html(
                    'College ID already exists. Use <a href="{}" target="_blank">Admin Panel</a>.',
                    ADMIN_URL))
                return redirect("university_reg_show")

            if User.objects.filter(username=uni.college_id).exists():
                messages.error(request, format_html(
                    'Username already exists. Use <a href="{}" target="_blank">Admin Panel</a>.',
                    ADMIN_URL))
                return redirect("university_reg_show")

            html_msg   = render_to_string("emails/approval_email.html", context)
            plain_msg  = strip_tags(html_msg)
            email_msg  = EmailMultiAlternatives(
                subject="SmartEduCare – University Account Approved 🎉",
                body=plain_msg,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[uni.university_email],
            )
            email_msg.attach_alternative(html_msg, "text/html")

            if os.path.isfile(REF_EXCEL_PATH):
                with open(REF_EXCEL_PATH, "rb") as f:
                    email_msg.attach("students_final_v8.xlsx", f.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                logger.warning("Reference Excel not found at %s", REF_EXCEL_PATH)

            email_msg.send(fail_silently=False)

            _save_credentials_to_excel(
                college_id=str(uni.college_id),
                email=uni.university_email,
                password=raw_password,
                name=uni.name,
            )

            user = User.objects.create_user(
                username=uni.college_id, email=uni.university_email,
                password=raw_password, first_name=uni.name[:30],
            )
            ApprovedUniversityAccount.objects.create(
                user=user, college_id=uni.college_id, name=uni.name,
                university_email=uni.university_email, country=uni.country,
                state=uni.state, pincode=uni.pincode,
                contact_number=uni.contact_number, website=uni.website,
            )
            ApprovedUniversityCredential.objects.create(
                college_id=uni.college_id, name=uni.name,
                email=uni.university_email, password=raw_password,
            )
            uni.status = "Approved"
            uni.save(update_fields=["status"])

            messages.success(request,
                f"✅ '{uni.name}' approved. Credentials emailed to {uni.university_email} "
                f"and saved to university_logins.xlsx.")

    except Exception as e:
        logger.exception("Approval failed pk=%s", pk)
        messages.error(request, f"Approval failed: {e}")

    return redirect("university_reg_show")

def approved_show(request):
    import os
    import datetime
    import shutil
    from django.conf import settings

    upload_results = []
    rules_message  = None

    # ✅ BASE PATHS
    BASE_STATIC = os.path.join(settings.BASE_DIR, "static", "data", "uploads")

    PW_GEN_DIR = os.path.join(BASE_STATIC, "password_generated_files")
    BACKUP_DIR = os.path.join(BASE_STATIC, "student_and_parent_logins")

    os.makedirs(PW_GEN_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)

    if request.method == "POST":

        # ── A) Student Files Upload ─────────────────────────────
        files = request.FILES.getlist("files[]")

        if files:
            aishe_code = request.POST.get("rules_aishe_code", "").strip()

            # Get university name
            university_name = aishe_code
            if aishe_code:
                try:
                    cred = ApprovedUniversityCredential.objects.get(college_id=aishe_code)
                    university_name = cred.name.strip()
                except ApprovedUniversityCredential.DoesNotExist:
                    pass

            ok_count  = 0
            err_count = 0

            folder_label = _safe_dirname(f"{university_name}_{aishe_code}")
            target_folder = os.path.join(PW_GEN_DIR, folder_label)

            for file_obj in files:
                ext = os.path.splitext(file_obj.name)[1].lower()

                if ext not in (".xlsx", ".xls", ".csv"):
                    upload_results.append({
                        "filename": file_obj.name,
                        "error": f"Unsupported type '{ext}'.",
                    })
                    err_count += 1
                    continue

                # ✅ PROCESS FILE
                result = _process_single_file(file_obj, university_name, aishe_code)
                upload_results.append(result)

                if result.get("error"):
                    err_count += 1
                else:
                    ok_count += 1

                    # ✅ BACKUP LOGIC (REMOVE PREFIX)
                    try:
                        if os.path.exists(target_folder):

                            files_in_dir = [
                                os.path.join(target_folder, f)
                                for f in os.listdir(target_folder)
                                if f.lower().endswith((".xlsx", ".xls", ".csv"))
                            ]

                            if files_in_dir:
                                # ✅ pick latest processed file
                                latest_file = max(files_in_dir, key=os.path.getmtime)

                                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

                                original_name = os.path.basename(latest_file)

                                # ✅ REMOVE PREFIX
                                if original_name.startswith("password_generated_"):
                                    clean_name = original_name.replace("password_generated_", "", 1)
                                else:
                                    clean_name = original_name

                                name, ext = os.path.splitext(clean_name)

                                backup_file = f"{name}_{timestamp}{ext}"
                                backup_path = os.path.join(BACKUP_DIR, backup_file)

                                print("✅ Found processed file:", latest_file)
                                print("✅ Backup saving as:", backup_file)

                                shutil.copy2(latest_file, backup_path)

                            else:
                                print("⚠️ No processed files found in:", target_folder)

                        else:
                            print("⚠️ Folder not found:", target_folder)

                    except Exception as e:
                        print("❌ Backup failed:", e)

            if ok_count:
                messages.success(
                    request,
                    f"✅ {ok_count} file(s) processed + clean backup created"
                )

            if err_count:
                messages.error(
                    request,
                    f"❌ {err_count} file(s) failed."
                )

        # ── B) Rules File Upload ─────────────────────────────
        if request.FILES.get("rules_file"):
            rules_file = request.FILES["rules_file"]
            aishe_code = request.POST.get("rules_aishe_code", "").strip()

            _, ext = os.path.splitext(rules_file.name.lower())

            if ext not in [".pdf", ".doc", ".docx"]:
                rules_message = "❌ Only PDF, DOC, DOCX allowed."
            else:
                rules_dir = os.path.join(BASE_STATIC, "AISHE_CODE", aishe_code, "rules")
                os.makedirs(rules_dir, exist_ok=True)

                save_path = os.path.join(rules_dir, rules_file.name)

                with open(save_path, "wb+") as dest:
                    for chunk in rules_file.chunks():
                        dest.write(chunk)

                rules_message = f"✅ Rules file '{rules_file.name}' saved."

    approved_list = ApprovedUniversityCredential.objects.all()

    return render(request, "approved_show.html", {
        "approved_list":  approved_list,
        "upload_results": upload_results,
        "rules_message":  rules_message,
        "message":        None,
    })

# ═════════════════════════════════════════════════════════════════════════════
#  SEND CORRECTION EMAIL  (unchanged)
# ═════════════════════════════════════════════════════════════════════════════
def send_correction_email(request, pk):
    uni = get_object_or_404(UniversityRequest, pk=pk)
    if uni.status != "Pending":
        messages.warning(request, "Only pending requests can be rejected.")
        return redirect("university_reg_show")

    remark = request.GET.get("remark", "").strip()
    if not remark:
        messages.error(request, "Remarks are required.")
        return redirect("university_reg_show")

    clean_name = re.sub(r"\d+", "",
                        uni.university_email.split("@")[0]).capitalize()
    try:
        html_message = render_to_string("emails/correction_email.html",
                                        {"clean_name": clean_name, "remark": remark})
        from django.core.mail import send_mail
        send_mail(
            subject="SmartEduCare – Registration Rejected",
            message=strip_tags(html_message),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[uni.university_email],
            html_message=html_message,
            fail_silently=False,
        )
        with transaction.atomic():
            uni.status = "Rejected"
            uni.save(update_fields=["status"])
            uni.delete()
        messages.success(request, f"Rejection email sent to {uni.university_email}.")
    except Exception:
        logger.exception("Rejection email failed")
        messages.error(request, "Email failed.")

    return redirect("university_reg_show")


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT DATA SEND  (unchanged)
# ═════════════════════════════════════════════════════════════════════════════
def student_data_send(request):
    context = {}
    if request.method == "POST":
        email         = request.POST.get("email")
        uploaded_file = request.FILES.get("file")
        if not email or not uploaded_file:
            context["error"] = "Email and file are required."
            return render(request, "student_data_send.html", context)

        upload_dir = os.path.join(settings.BASE_DIR,
                                  "static", "data", "uploads", "password_generated_files")
        os.makedirs(upload_dir, exist_ok=True)
        fs        = FileSystemStorage(location=upload_dir)
        filename  = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)

        try:
            msg = EmailMessage(
                subject="Student Login Credentials – Confidential",
                body=(
                    "Dear Team,\n\nPlease find attached the student data file.\n\n"
                    "🔐 Username: Student_ID | Password: student_password (in file)\n\n"
                    "⚠️ Strictly confidential.\n\nRegards,\nSmartEduCare Team"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            msg.attach_file(file_path)
            msg.send()
            context["message"] = "File sent successfully to the provided email."
        except Exception as e:
            context["error"] = f"Failed to send email: {e}"

    return render(request, "student_data_send.html", context)

import fitz
# ═════════════════════════════════════════════════════════════════════════════
#  ADMIN STUDENT DETAILS  (unchanged)
# ═════════════════════════════════════════════════════════════════════════════
def admin_student_details_show(request):
    students = student_login_valid.objects.all().order_by("-created_at")
    return render(request, "admin_student_details_show.html", {"students": students})
import os
import io
import base64
import json
import difflib
import tempfile
import pandas as pd
from django.conf import settings
from django.shortcuts import render
from django.http import JsonResponse

# ─────────────────────────────────────────────────────────────────
#  BACKUP DIR (still used for saving edits)
# ─────────────────────────────────────────────────────────────────
BACKUP_DIR = os.path.join(
    settings.BASE_DIR, "static", "data", "uploads", "student_and_parent_logins"
)

# ─────────────────────────────────────────────────────────────────
#  STRICT COLUMN POSITION + NAME RULES
#
#  University Excel format mandates:
#    • Column  1 (index 0)  →  StudentID   (case-sensitive, position 1)
#    • Column 10 (index 9)  →  PAN_CARD    (case-sensitive, position 10)
#    • Column 11 (index 10) →  AISHE_Code  (case-sensitive, position 11)
# ─────────────────────────────────────────────────────────────────

FUZZY_THRESHOLD = 0.60

COL_STUDENTID_IDX = 0
COL_PANCARD_IDX   = 9
COL_AISHE_IDX     = 10

CANONICAL_STUDENTID = "StudentID"
CANONICAL_PANCARD   = "PAN_CARD"
CANONICAL_AISHE     = "AISHE_Code"

STUDENTID_ALIASES = [
    "student id",   "student_id",   "studentid",    "stud id",
    "stud_id",      "roll",         "rollno",        "roll no",
    "roll_no",      "roll number",  "roll_number",   "regno",
    "reg no",       "reg_no",       "regid",         "reg id",
    "registration no","registrationno","enrollment no","enrollmentno",
    "enroll id",    "id number",    "student no",    "student number",
    "id",           "s.no",         "sno",           "sr no",
]

PAN_ALIASES = [
    "pan card",    "pan_card",     "pancard",      "pan no",
    "pan_no",      "panno",        "pan number",   "pan_number",
    "pannumber",   "pan id",       "panid",        "pan card no",
    "pan card number", "pan details", "pan doc",   "it pan",
    "income tax id","tax id",      "taxid",        "pan",
]

AISHE_ALIASES = [
    "aishe code",  "aishe_code",   "aishecode",    "aishe id",
    "aishe_id",    "aisheid",      "aishe no",     "aishe number",
    "institute code","institutecode","college code","collegecode",
    "inst code",   "instcode",     "college id",   "collegeid",
    "university code","universitycode","campus code","school code",
]

POSITION_TARGETS = {
    COL_STUDENTID_IDX: (CANONICAL_STUDENTID, STUDENTID_ALIASES),
    COL_PANCARD_IDX:   (CANONICAL_PANCARD,   PAN_ALIASES),
    COL_AISHE_IDX:     (CANONICAL_AISHE,     AISHE_ALIASES),
}


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    return (
        str(s).strip().lower()
        .replace("_", " ").replace("-", " ").replace(".", " ").strip()
    )


def _fuzzy_match(col_name: str, aliases: list, threshold: float = FUZZY_THRESHOLD) -> float:
    norm = _norm(col_name)
    best = 0.0
    for alias in aliases:
        r = difflib.SequenceMatcher(None, norm, alias).ratio()
        if len(norm) >= 5 and len(alias) >= 5 and (norm in alias or alias in norm):
            r = max(r, 0.72)
        if r > best:
            best = r
    return round(best, 3)


def _detect_header_row_from_df(raw: pd.DataFrame) -> int:
    """Detect header row from a raw (header=None) DataFrame."""
    best_row   = 0
    best_score = -1
    all_aliases = STUDENTID_ALIASES + PAN_ALIASES + AISHE_ALIASES

    for row_idx in range(min(15, len(raw))):
        vals = [
            str(v).strip()
            for v in raw.iloc[row_idx]
            if pd.notna(v) and str(v).strip() not in ("", "nan")
        ]
        non_unnamed = [v for v in vals if not v.startswith("Unnamed")]
        if len(non_unnamed) < 2:
            continue

        score = 0
        for cell in non_unnamed[:25]:
            n = _norm(cell)
            if n in STUDENTID_ALIASES:  score += 3; continue
            if n in PAN_ALIASES:        score += 3; continue
            if n in AISHE_ALIASES:      score += 3; continue
            if _fuzzy_match(cell, all_aliases) >= FUZZY_THRESHOLD:
                score += 1

        if score > best_score:
            best_score = score
            best_row   = row_idx

    return best_row


def _df_to_b64(df: pd.DataFrame, filename: str) -> str:
    """Serialize a DataFrame back to bytes (xlsx or csv) and return base64."""
    buf = io.BytesIO()
    if filename.lower().endswith(".csv"):
        csv_str = df.to_csv(index=False)
        buf.write(csv_str.encode("utf-8"))
    else:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _read_df_from_upload(file_obj, filename: str, header_row: int = None):
    """
    Read a Django InMemoryUploadedFile / TemporaryUploadedFile into a DataFrame.
    Returns (raw_df_no_header, used_header_row, df_with_header).
    """
    is_csv = filename.lower().endswith(".csv")
    file_obj.seek(0)
    raw_bytes = file_obj.read()
    file_obj.seek(0)

    buf = io.BytesIO(raw_bytes)

    # Read raw (no header) for detection
    buf.seek(0)
    try:
        raw = (
            pd.read_csv(buf, header=None, nrows=15, dtype=str)
            if is_csv
            else pd.read_excel(buf, header=None, nrows=15, dtype=str)
        )
    except Exception:
        raw = pd.DataFrame()

    detected_row = _detect_header_row_from_df(raw)
    used_header_row = header_row if header_row is not None else detected_row

    buf.seek(0)
    try:
        df = (
            pd.read_csv(buf, header=used_header_row, dtype=str)
            if is_csv
            else pd.read_excel(buf, header=used_header_row, dtype=str)
        )
        df = df.fillna("")
    except Exception as ex:
        raise ex

    return detected_row, used_header_row, df


def _process_df(df: pd.DataFrame, filename: str, detected_row: int, used_header_row: int) -> dict:
    """
    Apply position-aware fuzzy rename logic to an already-loaded DataFrame.
    Returns result dict (same shape as original _process_file).
    """
    original_cols = list(df.columns)
    rename_map    = {}
    col_results   = []
    not_found     = []

    for pos_idx, (canonical, aliases) in POSITION_TARGETS.items():
        pos_label = pos_idx + 1

        if pos_idx >= len(original_cols):
            not_found.append({
                "canonical":  canonical,
                "position":   pos_label,
                "reason":     f"File has only {len(original_cols)} column(s); "
                              f"column {pos_label} does not exist.",
                "suggestion": f"Open the file and ensure column {pos_label} "
                              f"contains the {canonical} header.",
            })
            continue

        col_name = original_cols[pos_idx]
        score    = _fuzzy_match(col_name, aliases)

        if score >= FUZZY_THRESHOLD:
            if col_name == canonical:
                action = "keep"
            else:
                action = "rename"
                rename_map[col_name] = canonical

            col_results.append({
                "original": col_name,
                "standard": canonical,
                "score":    score,
                "action":   action,
                "position": pos_label,
            })
        else:
            not_found.append({
                "canonical":  canonical,
                "position":   pos_label,
                "found_name": col_name,
                "score":      score,
                "reason":     (
                    f"Column {pos_label} is '{col_name}' "
                    f"(match score {round(score*100)}% < 60%). "
                    f"Expected something like '{canonical}'."
                ),
                "suggestion": (
                    f"Please rename column {pos_label} to exactly '{canonical}' (case-sensitive)."
                ),
            })
            col_results.append({
                "original": col_name,
                "standard": canonical,
                "score":    score,
                "action":   "not_found",
                "position": pos_label,
            })

    # Mark all non-target columns as skip
    target_positions = set(POSITION_TARGETS.keys())
    for i, col in enumerate(original_cols):
        if i not in target_positions:
            col_results.append({
                "original": col,
                "standard": None,
                "score":    0.0,
                "action":   "skip",
                "position": i + 1,
            })

    # Apply renames
    if rename_map:
        df = df.rename(columns=rename_map)

    preview_headers = list(df.columns)
    preview_rows = [
        [str(v) if v is not None else "" for v in row]
        for row in df.head(7).values.tolist()
    ]
    all_rows = [
        [str(v) if v is not None else "" for v in row]
        for row in df.values.tolist()
    ]

    return {
        "file":                filename,
        "status":              "ok",
        "header_row_detected": detected_row,
        "header_row_used":     used_header_row,
        "total_cols":          len(original_cols),
        "renamed":             sum(1 for r in col_results if r["action"] == "rename"),
        "already_ok":          sum(1 for r in col_results if r["action"] == "keep"),
        "skipped":             sum(1 for r in col_results if r["action"] == "skip"),
        "not_found_count":     len(not_found),
        "saved":               bool(rename_map),
        "columns":             col_results,
        "not_found":           not_found,
        "preview_headers":     preview_headers,
        "preview_rows":        preview_rows,
        "all_rows":            all_rows,
        "df":                  df,   # kept in memory for b64 serialisation
    }


# ─────────────────────────────────────────────────────────────────
#  MAIN VIEW
# ─────────────────────────────────────────────────────────────────
def admin_correct_data(request):
    os.makedirs(BACKUP_DIR, exist_ok=True)

    action = (
        request.GET.get("action") or request.POST.get("action") or ""
    ).strip()

    # ── POST: process_uploaded ─────────────────────────────────────
    # Accepts multipart files from anywhere on user's PC, processes them
    # in-memory, and returns JSON results + base64-encoded processed files.
    if request.method == "POST" and action == "process_uploaded":
        files = request.FILES.getlist("files")
        if not files:
            return JsonResponse({"error": "No files uploaded."}, status=400)

        raw_hr      = request.POST.get("header_row", "").strip()
        user_header_row = max(0, int(raw_hr)) if raw_hr.isdigit() else None

        results       = []
        total_renamed = 0

        for uploaded in files:
            fname = uploaded.name
            # Validate extension
            if not fname.lower().endswith((".xlsx", ".xls", ".csv")):
                results.append({
                    "file":   fname,
                    "status": "error",
                    "error":  "Unsupported file type. Only .xlsx, .xls, .csv allowed.",
                    "columns": [], "preview_headers": [], "preview_rows": [],
                    "all_rows": [], "not_found": [], "not_found_count": 0,
                    "renamed": 0, "already_ok": 0, "total_cols": 0,
                    "header_row_detected": 0, "header_row_used": 0, "saved": False,
                })
                continue

            try:
                detected_row, used_header_row, df = _read_df_from_upload(
                    uploaded, fname, user_header_row
                )
                res = _process_df(df, fname, detected_row, used_header_row)

                # Serialize processed (renamed) df to base64 for download
                processed_df = res.pop("df")
                res["processed_b64"] = _df_to_b64(processed_df, fname)

                total_renamed += res.get("renamed", 0)
                results.append(res)

            except Exception as ex:
                results.append({
                    "file":   fname,
                    "status": "error",
                    "error":  str(ex),
                    "columns": [], "preview_headers": [], "preview_rows": [],
                    "all_rows": [], "not_found": [], "not_found_count": 0,
                    "renamed": 0, "already_ok": 0, "total_cols": 0,
                    "header_row_detected": 0, "header_row_used": 0, "saved": False,
                })

        return JsonResponse({
            "files":           results,
            "total_files":     len(results),
            "ok_files":        sum(1 for r in results if r.get("status") == "ok"),
            "error_files":     sum(1 for r in results if r.get("status") != "ok"),
            "total_renamed":   total_renamed,
            "header_row_used": user_header_row,
        })

    # ── POST: apply_edit ──────────────────────────────────────────
    # Receives headers + all_rows as JSON, applies changes, returns new base64.
    if request.method == "POST" and action == "apply_edit":
        filename  = request.POST.get("filename", "").strip()
        roll_no   = request.POST.get("roll_no",   "").strip()
        changes   = request.POST.get("changes",   "{}")
        all_rows  = request.POST.get("all_rows",  "[]")
        headers   = request.POST.get("headers",   "[]")

        try:
            changes_dict  = json.loads(changes)
            rows_list     = json.loads(all_rows)
            headers_list  = json.loads(headers)
        except Exception as ex:
            return JsonResponse({"error": "Invalid JSON payload: " + str(ex)}, status=400)

        if not filename:
            return JsonResponse({"error": "No filename."}, status=400)

        try:
            df = pd.DataFrame(rows_list, columns=headers_list)

            # Apply changes for matching rows
            id_col = headers_list[0] if headers_list else None
            if id_col:
                mask = df[id_col].astype(str).str.strip() == roll_no.strip()
                for col, new_val in changes_dict.items():
                    if col in df.columns:
                        df.loc[mask, col] = new_val.strip()

            # Save backup to disk
            os.makedirs(BACKUP_DIR, exist_ok=True)
            backup_path = os.path.join(BACKUP_DIR, filename)
            if filename.lower().endswith(".csv"):
                df.to_csv(backup_path, index=False)
            else:
                df.to_excel(backup_path, index=False)

            processed_b64 = _df_to_b64(df, filename)

            return JsonResponse({
                "status":        "ok",
                "processed_b64": processed_b64,
                "updated_cols":  list(changes_dict.keys()),
            })

        except Exception as ex:
            return JsonResponse({"error": str(ex)}, status=500)

    # ── GET: page render ──────────────────────────────────────────
    # All file selection/upload is now done client-side via the file input.
    # The page render no longer needs folder listing or file state.
    header_row_val = (
        request.POST.get("header_row") or request.GET.get("header_row") or ""
    ).strip()
    header_row_val = header_row_val if header_row_val.isdigit() else ""

    return render(request, "admin_correct_data.html", {
        "header_row_val": header_row_val,
        "col_positions": {
            "studentid": COL_STUDENTID_IDX + 1,
            "pan_card":  COL_PANCARD_IDX   + 1,
            "aishe":     COL_AISHE_IDX     + 1,
        },
    })

# ============================================================
#  views.py  —  College Data Upload
#  Save path: static/data/uploads/College_rules/<college_name>/
# ============================================================

# ─────────────────────────────────────────────────────────────
# REQUIRED IMPORTS
# ─────────────────────────────────────────────────────────────
import os
import json
import pathlib
import textwrap
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings

try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

try:
    import fitz
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from docx import Document as DocxDocument
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False


# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────
GROQ_KEY3 = os.getenv("GROQ_API_KEY3", "")   # Whisper
GROQ_KEY5 = os.getenv("GROQ_API_KEY5", "")   # Chat

CHAT_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"

COLLEGE_ROOT = pathlib.Path(
    getattr(settings, "COLLEGE_RULES_ROOT",
            "static/data/uploads/College_rules")
)

ALLOWED_EXT = {".pdf", ".doc", ".docx", ".txt"}

MAX_CONTEXT_CHARS = 10000
MAX_HISTORY_TURNS = 10


# ─────────────────────────────────────────────────────────────
# SAFE PATH
# ─────────────────────────────────────────────────────────────
def _safe_path(college, filename=""):
    base = COLLEGE_ROOT.resolve()
    target = (COLLEGE_ROOT / college / filename).resolve() if filename else (COLLEGE_ROOT / college).resolve()

    if not str(target).startswith(str(base)):
        raise ValueError("Path traversal blocked")

    return target


# ─────────────────────────────────────────────────────────────
# TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────
def _extract_pdf(path):
    doc = fitz.open(str(path))
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    return text.strip()


def _extract_docx(path):
    doc = DocxDocument(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_txt(path):
    return path.read_text(encoding="utf-8", errors="ignore")


def _extract_text(path):
    ext = path.suffix.lower()

    if ext == ".pdf":
        return _extract_pdf(path)
    elif ext in [".doc", ".docx"]:
        return _extract_docx(path)
    elif ext == ".txt":
        return _extract_txt(path)

    return "Unsupported file"


# ─────────────────────────────────────────────────────────────
# STRICT PROMPT (NO HALLUCINATION)
# ─────────────────────────────────────────────────────────────
def _build_system_prompt(college, file_name, context):
    return f"""
You are a STRICT document-based AI assistant.

RULES:
- Answer ONLY from given document
- DO NOT use external knowledge
- DO NOT guess
- If not found say:
  "This information is not available in the selected document."

College: {college}
Document: {file_name}

DOCUMENT:
{context[:MAX_CONTEXT_CHARS]}
"""


# ─────────────────────────────────────────────────────────────
# GROQ CHAT
# ─────────────────────────────────────────────────────────────
def _groq_chat(system, messages):
    client = Groq(api_key=GROQ_KEY5)

    response = client.chat.completions.create(
        model=CHAT_MODEL,
        messages=[{"role": "system", "content": system}] + messages,
        temperature=0.1,
        max_tokens=1024
    )

    return response.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────
# WHISPER (AUDIO → TEXT)
# ─────────────────────────────────────────────────────────────
def _speech_to_text(audio_file_path):
    client = Groq(api_key=GROQ_KEY3)

    with open(audio_file_path, "rb") as f:
        transcript = client.audio.transcriptions.create(
            file=f,
            model="whisper-large-v3"
        )

    return transcript.text


# ─────────────────────────────────────────────────────────────
# MAIN VIEW
# ─────────────────────────────────────────────────────────────
@csrf_exempt
def know_about_your_college(request):

    action = request.GET.get("action", "")

    # ─── PAGE LOAD ───
    if not action:
        return render(request, "know_about_your_college.html")

    # ─── LIST COLLEGES ───
    if action == "list_colleges":
        COLLEGE_ROOT.mkdir(parents=True, exist_ok=True)
        colleges = [p.name for p in COLLEGE_ROOT.iterdir() if p.is_dir()]
        return JsonResponse({"colleges": colleges})

    # ─── LIST FILES ───
    if action == "list_files":
        college = request.GET.get("college")

        folder = _safe_path(college)

        files = [f.name for f in folder.iterdir()
                 if f.is_file() and f.suffix.lower() in ALLOWED_EXT]

        return JsonResponse({"files": files})

    # ─── EXTRACT TEXT ───
    if action == "extract":
        college = request.GET.get("college")
        file = request.GET.get("file")

        path = _safe_path(college, file)

        text = _extract_text(path)

        return JsonResponse({
            "text": text[:MAX_CONTEXT_CHARS],
            "full_length": len(text)
        })

    # ─── CHAT (STRICT RAG) ───
    if action == "chat" and request.method == "POST":

        data = json.loads(request.body)

        message = data.get("message", "")
        context = data.get("context", "")
        college = data.get("college", "")
        file_name = data.get("file", "")
        history = data.get("history", [])[-MAX_HISTORY_TURNS:]

        # 🚫 BLOCK IF NO DOCUMENT
        if not context:
            return JsonResponse({
                "reply": "⚠️ Please select a document first."
            })

        system = _build_system_prompt(college, file_name, context)

        messages = history + [{"role": "user", "content": message}]

        reply = _groq_chat(system, messages)

        return JsonResponse({"reply": reply})

    # ─── SPEECH INPUT (OPTIONAL) ───
    if action == "speech" and request.method == "POST":

        audio_file = request.FILES.get("audio")

        temp_path = "temp_audio.wav"

        with open(temp_path, "wb+") as f:
            for chunk in audio_file.chunks():
                f.write(chunk)

        text = _speech_to_text(temp_path)

        os.remove(temp_path)

        return JsonResponse({"text": text})

    return JsonResponse({"error": "Invalid action"})

from django.shortcuts import render, redirect
from django.conf import settings
import os

def college_data_upload(request):
    if request.method == "POST":
        college = request.POST.get("college")
        file = request.FILES.get("file")

        if college and file:
            folder = os.path.join(
                settings.BASE_DIR,
                "static/data/uploads/College_rules",
                college
            )
            os.makedirs(folder, exist_ok=True)

            file_path = os.path.join(folder, file.name)

            with open(file_path, "wb+") as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

        return redirect("college_data_upload")

    return render(request, "college_data_upload.html")